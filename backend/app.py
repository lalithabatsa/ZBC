import os
import re
import glob
import pandas as pd
from fastapi import FastAPI, HTTPException, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Dict, Any
from datetime import datetime
import openpyxl

app = FastAPI()

PRICING_FOLDER = "../data/scraped_data"
if not os.path.exists(PRICING_FOLDER):
    os.makedirs(PRICING_FOLDER, exist_ok=True)

def get_latest_file():
    files = [f for f in os.listdir(PRICING_FOLDER) if f.endswith('.xlsx')]
    if not files:
        return None
    
    # Sort by date in filename (format: DD-MM-YYYY-...)
    def extract_date(filename):
        match = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
        if match:
            return datetime.strptime(match.group(0), '%d-%m-%Y')
        return datetime.min

    latest = max(files, key=extract_date)
    return os.path.join(PRICING_FOLDER, latest)

def parse_excel(file_path):
    filename = os.path.basename(file_path)
    date_match = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
    if date_match:
        day, month, year = date_match.groups()
        months = ["January", "February", "March", "April", "May", "June", 
                  "July", "August", "September", "October", "November", "December"]
        filename_date = f"{day} {months[int(month)-1]} {year}"
    else:
        filename_date = "Unknown"
    
    print(f"Parsing {filename}")
    xl = pd.ExcelFile(file_path)
    sheet_name = next((s for s in xl.sheet_names if s.lower() == "ex-works"), xl.sheet_names[0])
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    records = []
    i = 0
    while i < len(df):
        row_values = [str(val).strip() if pd.notna(val) else "" for val in df.iloc[i]]
        
        # Look for the main header row (Sr.No., Pricing Zone, State)
        if "Sr.No." in row_values and "Pricing Zone" in row_values and "State" in row_values:
            zone_idx = row_values.index("Pricing Zone")
            state_idx = row_values.index("State")
            sr_idx = row_values.index("Sr.No.")
            
            # Metadata search (Upwards)
            meta = {"date": "Unknown", "ref": "Unknown", "location": "Unknown", "annexure": "Unknown"}
            for k in range(max(0, i-6), i):
                prev_row_values = [str(val) for val in df.iloc[k] if pd.notna(val)]
                prev_row = " ".join(prev_row_values)
                
                date_match = re.search(r'Date\s*[:\-]?\s*([A-Za-z0-9\s,]+)(?=Ref|$)', prev_row, re.I)
                if date_match: meta["date"] = date_match.group(1).strip()
                
                ref_match = re.search(r'Ref\s*[:\-]?\s*([A-Z0-9\/\-\s]+)', prev_row, re.I)
                if ref_match: meta["ref"] = ref_match.group(1).strip()
                
                if "Domestic Prices for" in prev_row or "Prices for" in prev_row:
                    loc_match = re.search(r'(?:Domestic\s+)?Prices for\s+(.+?)(?:\s+(HDPE|LLDPE|LDPE|PE|NYLON|TIE|RESIN)|\.|\s*Annexure|$)', prev_row, re.I)
                    if loc_match:
                        location_str = loc_match.group(1).strip()
                        meta["location"] = location_str.split()[0] if location_str else "Unknown"
                        if len(loc_match.groups()) > 1 and loc_match.group(2):
                            meta["table_material"] = loc_match.group(2).upper()
                
                if "Annexure" in prev_row:
                    ann_match = re.search(r'Annexure\s*[:\-]?\s*([A-Z0-9\s\-]+)', prev_row, re.I)
                    if ann_match: meta["annexure"] = ann_match.group(1).strip()

            if meta.get("table_material") == "HDPE":
                i += 1
                continue

            header = row_values
            grades = [h for h in header[state_idx+1:] if h and h != "None" and "Unnamed" not in h]
            
            # Parse rows below the header
            j = i + 1
            while j < len(df):
                sub_row = df.iloc[j]
                sr_no = str(sub_row[sr_idx]).strip() if pd.notna(sub_row[sr_idx]) else ""
                zone = str(sub_row[zone_idx]).strip() if pd.notna(sub_row[zone_idx]) else ""
                state = str(sub_row[state_idx]).strip() if pd.notna(sub_row[state_idx]) else ""
                
                if not zone or zone.lower() == "none" or "total" in zone.lower() or not sr_no.isdigit():
                    break
                
                for k, grade in enumerate(grades):
                    if state_idx + 1 + k >= len(sub_row): continue
                    price_val = sub_row[state_idx + 1 + k]
                    if pd.notna(price_val) and str(price_val).replace('.', '', 1).isdigit():
                        records.append({
                            "material": meta["location"] + " " + (os.path.basename(os.path.dirname(file_path))),
                            "location": meta["location"],
                            "state": state,
                            "zone": zone,
                            "grade": grade,
                            "price": float(price_val),
                            "table_material": meta.get("table_material"),
                            "meta": {
                                **meta,
                                "filename_date": filename_date,
                                "source_file": filename,
                                "sheet": sheet_name,
                                "cell_ref": f"{chr(65 + state_idx + 1 + k)}{j + 1}"
                            }
                        })
                j += 1
            i = j # Move main loop index forward
        else:
            i += 1
            
    return records

# Global cache to avoid heavy re-parsing
_cached_data = None
_last_load_time = None

def load_all_data():
    global _cached_data, _last_load_time
    
    # Check cache (1 min validity)
    if _cached_data and _last_load_time and (datetime.now() - _last_load_time).seconds < 60:
        return _cached_data

    # Use a dictionary to track unique files by basename to avoid redundant parsing
    all_files = glob.glob(os.path.join(PRICING_FOLDER, "**/*.xlsx"), recursive=True)
    unique_files = {}
    for f in all_files:
        basename = os.path.basename(f)
        if basename.startswith("~$") or "Reliance" not in basename:
            continue
        # If we have multiple copies of the same file, we only need one
        if basename not in unique_files:
            unique_files[basename] = f

    def get_file_date(filepath):
        filename = os.path.basename(filepath)
        match = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
        return datetime.strptime(match.group(0), '%d-%m-%Y') if match else datetime.min

    sorted_files = sorted(unique_files.values(), key=get_file_date, reverse=True)

    all_records = []
    # Dictionary to keep only the latest record for each unique combination
    latest_lookup = {}

    for f in sorted_files:
        try:
            data = parse_excel(f)
            folder_name = os.path.basename(os.path.dirname(f))
            for d in data:
                t_mat = d.get("table_material")
                d["folder_key"] = t_mat if t_mat in ["LDPE", "LLDPE"] else folder_name
                
                # Create a unique key for this specific price point
                # We ignore the filename/date in the key to find duplicates across files
                key = (d["material"], d["location"], d["state"], d["zone"], d["grade"])
                
                if key not in latest_lookup:
                    latest_lookup[key] = d
                    all_records.append(d)
        except Exception as e:
            print(f"Error parsing {f}: {e}")
            
    _cached_data = all_records
    _last_load_time = datetime.now()
    return all_records

from fastapi import BackgroundTasks

def run_scraper():
    print("Background scraper started...")
    try:
        from Scraper_2 import URLS, process_polymer
        for polymer, url in URLS.items():
            process_polymer(polymer, url)
        print("Background scraper finished.")
        # Invalidate cache after scraping new data
        global _cached_data
        _cached_data = None
    except Exception as e:
        print(f"Background scraper failed: {e}")

@app.post("/api/scrape")
async def trigger_scrape(background_tasks: BackgroundTasks):
    background_tasks.add_task(run_scraper)
    return {"status": "pending", "message": "Scraping started in background. Refresh data in a few moments."}

@app.get("/api/data")
async def get_data():
    try:
        data = load_all_data()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/data/material/{material}")
async def get_material_data(material: str):
    try:
        all_data = load_all_data()
        # Filter by material name or folder key
        filtered = [
            d for d in all_data 
            if material.upper() in d["material"].upper() or material.upper() in d.get("folder_key", "").upper()
        ]
        return filtered
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class PriceUpdate(BaseModel):
    filename: str
    sheet: str
    cell_ref: str
    new_price: float

@app.post("/api/update-price")
async def update_price(update: PriceUpdate):
    file_path = os.path.join(PRICING_FOLDER, update.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        if update.sheet not in wb.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet {update.sheet} not found")
        
        ws = wb[update.sheet]
        ws[update.cell_ref] = update.new_price
        wb.save(file_path)
        return {"status": "success", "message": f"Updated {update.cell_ref} to {update.new_price}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
async def read_index():
    return FileResponse('../frontend/index.html')

# Serve other static files
app.mount("/", StaticFiles(directory="../frontend", html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
